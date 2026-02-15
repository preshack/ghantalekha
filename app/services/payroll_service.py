import csv
import io
import logging
from datetime import datetime, timezone
from calendar import monthrange

from sqlalchemy import func, extract

from app.extensions import db
from app.models.employee import Employee
from app.models.attendance import Attendance

logger = logging.getLogger(__name__)


def get_month_range(year, month):
    """Return start and end datetime for a given month (UTC)."""
    start = datetime(year, month, 1, tzinfo=timezone.utc)
    last_day = monthrange(year, month)[1]
    end = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)
    return start, end


def get_monthly_hours(employee_id, year, month):
    """Calculate total hours and overtime for an employee in a given month.

    Returns:
        dict: {total_minutes, total_hours, regular_hours, overtime_hours}
    """
    from flask import current_app
    threshold = current_app.config.get('OVERTIME_MONTHLY_THRESHOLD', 160)

    start, end = get_month_range(year, month)

    total_minutes = (db.session.query(func.coalesce(func.sum(Attendance.work_duration_minutes), 0))
                     .filter(
                         Attendance.employee_id == employee_id,
                         Attendance.clock_in >= start,
                         Attendance.clock_in <= end,
                         Attendance.work_duration_minutes.isnot(None)
                     )
                     .scalar()) or 0

    total_hours = round(total_minutes / 60, 2)
    regular_hours = min(total_hours, threshold)
    overtime_hours = max(0, total_hours - threshold)

    return {
        'total_minutes': total_minutes,
        'total_hours': total_hours,
        'regular_hours': regular_hours,
        'overtime_hours': overtime_hours,
    }


def get_today_hours(employee_id):
    """Calculate total hours worked today for an employee."""
    now = datetime.now(timezone.utc)
    start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)

    total_minutes = (db.session.query(func.coalesce(func.sum(Attendance.work_duration_minutes), 0))
                     .filter(
                         Attendance.employee_id == employee_id,
                         Attendance.clock_in >= start_of_day,
                         Attendance.work_duration_minutes.isnot(None)
                     )
                     .scalar()) or 0

    return round(total_minutes / 60, 2)


def get_all_employees_monthly_summary(year, month):
    """Generate monthly summary for all active employees.

    Returns:
        list of dicts: [{employee, total_hours, regular_hours, overtime_hours, regular_pay, overtime_pay, total_pay}]
    """
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
    summaries = []

    for emp in employees:
        hours = get_monthly_hours(emp.id, year, month)
        hourly_rate = float(emp.hourly_rate)

        regular_pay = round(hours['regular_hours'] * hourly_rate, 2)
        overtime_pay = round(hours['overtime_hours'] * hourly_rate * 1.5, 2)  # 1.5x for overtime
        total_pay = round(regular_pay + overtime_pay, 2)

        summaries.append({
            'employee': emp,
            'employee_id': emp.id,
            'employee_name': emp.name,
            'email': emp.email,
            'hourly_rate': hourly_rate,
            'total_hours': hours['total_hours'],
            'regular_hours': hours['regular_hours'],
            'overtime_hours': hours['overtime_hours'],
            'regular_pay': regular_pay,
            'overtime_pay': overtime_pay,
            'total_pay': total_pay,
        })

    return summaries


def get_dashboard_metrics(year, month):
    """Get dashboard summary metrics.

    Returns:
        dict with keys: active_count, total_hours_today, monthly_summaries,
                        total_monthly_hours, total_payroll, total_overtime_pay
    """
    from app.services.attendance_service import get_active_employees_count

    summaries = get_all_employees_monthly_summary(year, month)

    # Total hours today across all employees
    now = datetime.now(timezone.utc)
    start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
    total_today_minutes = (db.session.query(func.coalesce(func.sum(Attendance.work_duration_minutes), 0))
                           .filter(
                               Attendance.clock_in >= start_of_day,
                               Attendance.work_duration_minutes.isnot(None)
                           )
                           .scalar()) or 0

    total_monthly_hours = sum(s['total_hours'] for s in summaries)
    total_payroll = sum(s['total_pay'] for s in summaries)
    total_overtime_pay = sum(s['overtime_pay'] for s in summaries)

    return {
        'active_count': get_active_employees_count(),
        'total_hours_today': round(total_today_minutes / 60, 2),
        'monthly_summaries': summaries,
        'total_monthly_hours': round(total_monthly_hours, 2),
        'total_payroll': round(total_payroll, 2),
        'total_overtime_pay': round(total_overtime_pay, 2),
    }


def get_employee_monthly_log(employee_id, year, month):
    """Get all attendance records for an employee in a given month."""
    start, end = get_month_range(year, month)

    records = (Attendance.query
               .filter(
                   Attendance.employee_id == employee_id,
                   Attendance.clock_in >= start,
                   Attendance.clock_in <= end,
               )
               .order_by(Attendance.clock_in.desc())
               .all())

    return records


def generate_payroll_csv(year, month):
    """Generate a CSV file for payroll data.

    Returns:
        io.StringIO: CSV content ready for download or email attachment.
    """
    summaries = get_all_employees_monthly_summary(year, month)

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        'Employee ID', 'Employee Name', 'Email', 'Hourly Rate',
        'Total Hours', 'Regular Hours', 'Overtime Hours',
        'Regular Pay', 'Overtime Pay (1.5x)', 'Total Pay'
    ])

    for s in summaries:
        writer.writerow([
            s['employee_id'], s['employee_name'], s['email'],
            f"${s['hourly_rate']:.2f}",
            f"{s['total_hours']:.2f}", f"{s['regular_hours']:.2f}", f"{s['overtime_hours']:.2f}",
            f"${s['regular_pay']:.2f}", f"${s['overtime_pay']:.2f}", f"${s['total_pay']:.2f}",
        ])

    # Totals row
    writer.writerow([])
    writer.writerow([
        '', '', '', 'TOTALS',
        f"{sum(s['total_hours'] for s in summaries):.2f}",
        f"{sum(s['regular_hours'] for s in summaries):.2f}",
        f"{sum(s['overtime_hours'] for s in summaries):.2f}",
        f"${sum(s['regular_pay'] for s in summaries):.2f}",
        f"${sum(s['overtime_pay'] for s in summaries):.2f}",
        f"${sum(s['total_pay'] for s in summaries):.2f}",
    ])

    output.seek(0)
    return output


def generate_payroll_excel(year, month):
    """Generate an Excel file for payroll data.

    Returns:
        io.BytesIO: Excel content ready for download.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    summaries = get_all_employees_monthly_summary(year, month)
    wb = Workbook()
    ws = wb.active
    ws.title = f'Payroll {year}-{month:02d}'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    currency_format = '$#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f'WorkClock Payroll Report — {year}-{month:02d}'
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    # Headers
    headers = [
        'Employee ID', 'Employee Name', 'Email', 'Hourly Rate',
        'Total Hours', 'Regular Hours', 'Overtime Hours',
        'Regular Pay', 'Overtime Pay (1.5x)', 'Total Pay'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data
    for row_idx, s in enumerate(summaries, 4):
        data = [
            s['employee_id'], s['employee_name'], s['email'],
            s['hourly_rate'], s['total_hours'], s['regular_hours'],
            s['overtime_hours'], s['regular_pay'], s['overtime_pay'], s['total_pay']
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col >= 4:
                cell.number_format = currency_format if col in (4, 8, 9, 10) else '0.00'

    # Totals row
    total_row = len(summaries) + 5
    ws.cell(row=total_row, column=4, value='TOTALS').font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(s['total_hours'] for s in summaries)).number_format = '0.00'
    ws.cell(row=total_row, column=6, value=sum(s['regular_hours'] for s in summaries)).number_format = '0.00'
    ws.cell(row=total_row, column=7, value=sum(s['overtime_hours'] for s in summaries)).number_format = '0.00'
    ws.cell(row=total_row, column=8, value=sum(s['regular_pay'] for s in summaries)).number_format = currency_format
    ws.cell(row=total_row, column=9, value=sum(s['overtime_pay'] for s in summaries)).number_format = currency_format
    ws.cell(row=total_row, column=10, value=sum(s['total_pay'] for s in summaries)).number_format = currency_format

    # Auto-width columns
    from openpyxl.utils import get_column_letter
    for col in range(1, 11):
        max_length = max(len(str(ws.cell(row=r, column=col).value or '')) for r in range(3, total_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = max(max_length + 2, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
